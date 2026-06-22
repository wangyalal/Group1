import ttkbootstrap as tb
from ttkbootstrap.constants import BOTH, YES
from tkinter import messagebox, ttk
import tkinter as tk
from datetime import date
from decimal import Decimal, InvalidOperation
from transactions import get_all_transactions
from database import get_db_connection
from timeframe import get_date_range, filter_transactions, PRESETS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import filedialog


def _get_transaction_id(tx_date, description, amount, tx_type):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id FROM transactions
            WHERE transaction_date = %s
              AND description      = %s
              AND ABS(amount)      = %s
              AND transaction_type = %s
            ORDER BY id DESC LIMIT 1;
        """, (tx_date, description, Decimal(str(abs(amount))), tx_type))
        row = cursor.fetchone()
        return row[0] if row else None
    except Exception as e:
        print(f"Lookup error: {e}")
        return None
    finally:
        cursor.close()
        conn.close()


def _delete_transaction(transaction_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM transactions WHERE id = %s;", (transaction_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Delete error: {e}")
        return False
    finally:
        cursor.close()
        conn.close()


def _update_transaction(transaction_id: int, tx_date, amount, description,
                        tx_type: str, category_name: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            UPDATE transactions
            SET transaction_date = %s,
                amount           = %s,
                description      = %s,
                transaction_type = %s,
                category         = %s
            WHERE id = %s;
        """, (tx_date, Decimal(str(amount)), description, tx_type, category_name, transaction_id))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"Update error: {e}")
        return False
    finally:
        cursor.close()
        conn.close()

def export_to_excel(transactions, summary, start, end):
    
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="transactions_export.xlsx",
        title="Export Transactions"
    )
    if not filepath:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Period label
    if start and end:
        period = f"{start.strftime('%b %d, %Y')} → {end.strftime('%b %d, %Y')}"
    else:
        period = "All Time"

    ws["A1"] = "Expense Tracker — Transaction Export"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Period: {period}"
    ws["A2"].font = Font(italic=True, size=10)

    # Header row
    headers = ["Date", "Description", "Category", "Type", "Amount"]
    header_fill   = PatternFill("solid", start_color="2962FF", end_color="2962FF")
    header_font   = Font(bold=True, color="FFFFFF")
    header_align  = Alignment(horizontal="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align

    # Data rows
    income_font  = Font(color="28A745")
    expense_font = Font(color="DC3545")

    for row_idx, row in enumerate(transactions, start=5):
        tx_date, desc, cat, amount, tx_type = row
        signed = amount if tx_type == "Income" else -abs(amount)

        ws.cell(row=row_idx, column=1, value=str(tx_date))
        ws.cell(row=row_idx, column=2, value=desc)
        ws.cell(row=row_idx, column=3, value=cat)
        ws.cell(row=row_idx, column=4, value=tx_type)

        amt_cell = ws.cell(row=row_idx, column=5, value=float(signed))
        amt_cell.number_format = '#,##0.00'
        amt_cell.font = income_font if tx_type == "Income" else expense_font

    # Net row
    net_row = len(transactions) + 5
    ws.cell(row=net_row, column=1, value=f"Net ({summary['count']} transactions)").font = Font(bold=True)
    net_cell = ws.cell(row=net_row, column=5, value=float(summary["net"]))
    net_cell.number_format = '#,##0.00'
    net_cell.font = Font(bold=True, color="28A745" if summary["net"] >= 0 else "DC3545")

    # Summary below
    sum_row = net_row + 2
    ws.cell(row=sum_row,     column=1, value="Total Income").font  = Font(bold=True)
    ws.cell(row=sum_row,     column=2, value=float(summary["total_income"])).number_format = '#,##0.00'
    ws.cell(row=sum_row + 1, column=1, value="Total Expenses").font = Font(bold=True)
    ws.cell(row=sum_row + 1, column=2, value=float(summary["total_expenses"])).number_format = '#,##0.00'
    ws.cell(row=sum_row + 2, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=sum_row + 2, column=2, value=summary["status"])

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    wb.save(filepath)
    return True
def get_transaction_ledger(preset="All Time", cycle_start_day=1,
                           custom_start=None, custom_end=None):
                               
    all_transactions = get_all_transactions()
    start, end       = get_date_range(preset, cycle_start_day, custom_start, custom_end)
    transactions     = filter_transactions(all_transactions, start, end)

    total_income   = sum(row[3] for row in transactions if row[4] == "Income")
    total_expenses = sum(abs(row[3]) for row in transactions if row[4] == "Expense")
    net            = total_income - total_expenses

    summary = {
        "total_income":   total_income,
        "total_expenses": total_expenses,
        "net":            net,
        "status":         "Under control" if net >= 0 else "Overspending",
        "count":          len(transactions),
        "start":          start,
        "end":            end,
    }
    return transactions, summary


class _CellEditor:

    EXPENSE_CATEGORIES = ["Food", "Rent", "Phone", "Utilities", "Commute", "Leisure", "Other"]
    INCOME_CATEGORIES  = ["Salary", "Bonus", "Other"]
    TYPE_OPTIONS       = ["Income", "Expense"]

    def __init__(self, tree: tb.Treeview, item_id: str, col_index: int,
                 current_value: str, on_commit):
        self._tree      = tree
        self._item_id   = item_id
        self._col_index = col_index
        self._on_commit = on_commit
        self._widget    = None

        col_id   = tree["columns"][col_index]
        bbox     = tree.bbox(item_id, column=col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        col_name = tree.heading(col_id)["text"]

        editor_font = ("Helvetica", 10)

        if col_name == "Type":
            self._widget = tb.Combobox(tree, values=self.TYPE_OPTIONS,
                                        state="readonly", font=editor_font)
            self._widget.set(current_value)
            self._widget.bind("<<ComboboxSelected>>", self._commit)

        elif col_name == "Category":
            values_col = tree["columns"][3]
            tx_type    = tree.set(item_id, values_col)
            cats = self.INCOME_CATEGORIES if tx_type == "Income" else self.EXPENSE_CATEGORIES
            self._widget = tb.Combobox(tree, values=cats,
                                        state="readonly", font=editor_font)
            self._widget.set(current_value)
            self._widget.bind("<<ComboboxSelected>>", self._commit)

        else:
            self._widget = tb.Entry(tree, font=editor_font)
            self._widget.insert(0, current_value)
            self._widget.select_range(0, "end") 
            self._widget.bind("<Return>", self._commit)
            self._widget.bind("<Escape>", self._cancel)
            self._widget.bind("<FocusOut>", self._commit)

        adjusted_height = h+6
        adjusted_y = y-3

        self._widget.place(x=x, y=adjusted_y, width=w, height=adjusted_height)
        self._widget.focus_set()

    def _commit(self, _event=None):
        if self._widget is None:
            return
        value = self._widget.get()
        self._widget.destroy()
        self._widget = None
        self._on_commit(self._item_id, self._col_index, value)

    def _cancel(self, _event=None):
        if self._widget:
            self._widget.destroy()
            self._widget = None



class LedgerFrame(tb.Frame):
  
    COLUMNS = [
        ("Date",        True),
        ("Description", True),
        ("Category",    True),
        ("Type",        True),
        ("Amount",      True),
    ]

    _BUTTON_PRESETS = [
        "All Time",
        "This Month",
        "Last Month",
        "Last 7 Days",
        "Last 30 Days",
        "This Year",
    ]

    def __init__(self, parent, app_reference = None, **kwargs):
        super().__init__(parent, **kwargs)
        self._tree            = None
        self._raw_rows        = []
        self._editor          = None
        self._preset          = "All Time"
        self._cycle_start_day = 1
        self._custom_start    = None
        self._custom_end      = None
        self._preset_buttons  = {}
        self._app_ref = app_reference
        self._build_controls()
        self._refresh()


    def set_custom_range(self, start: date, end: date):
        """
        Call this from outside (e.g. gui.py) to apply a custom date range.
        Switches the active preset to Custom and refreshes the table.
        """
        self._custom_start = start
        self._custom_end   = end
        self._set_preset("Custom")


    def _build_controls(self):
        bar = tb.Frame(self)
        bar.pack(padx=20, pady=(10, 4), anchor= "center")

        for preset in self._BUTTON_PRESETS:
            btn = tb.Button(
                bar, text=preset, width=12,
                bootstyle="primary" if preset == self._preset else "outline-dark",
                command=lambda p=preset: self._set_preset(p),
            )
            btn.pack(side="left", padx=(0, 4))
            self._preset_buttons[preset] = btn

    def _set_preset(self, preset: str):
        self._preset = preset
        for p, btn in self._preset_buttons.items():
            btn.configure(bootstyle="primary" if p == preset else "outline-secondary")
        self._refresh()

    def _on_cycle_change(self):
        try:
            self._cycle_start_day = int(self._day_var.get())
        except ValueError:
            self._cycle_start_day = 1
        self._refresh()

    def _refresh(self):
        children = self.winfo_children()
        for widget in children[1:]:
            widget.destroy()
        self._editor = None

        transactions, summary = get_transaction_ledger(
            preset=self._preset,
            cycle_start_day=self._cycle_start_day,
            custom_start=self._custom_start,
            custom_end=self._custom_end,
        )
        self._raw_rows = list(transactions)

        self._build_period_label(summary)
        self._build_summary_cards(summary)
        self._build_action_bar()
        self._build_table(transactions, summary)


    def _build_period_label(self, summary):
        start, end = summary["start"], summary["end"]
        if start and end:
            text = f"{start.strftime('%b %d, %Y')}  →  {end.strftime('%b %d, %Y')}"
        else:
            text = "All Time"

        tb.Label(self, text=text, font=("Helvetica", 9, "bold"), foreground= "#000000",
                 bootstyle="secondary").pack(anchor="w", padx=22, pady=(2, 0)) 


    def _build_summary_cards(self, summary):
        card_row = tb.Frame(self)
        card_row.pack(fill="x", padx=20, pady=(6, 8))

        cards = [
            ("Total Income",   f"+${summary['total_income']:,.2f}",   "success-inverse"),
            ("Total Expenses", f"-${summary['total_expenses']:,.2f}", "danger-inverse"),
            (
                "Net Balance:",
                f"{'+' if summary['net'] >= 0 else ''}${summary['net']:,.2f}",
                "success" if summary["net"] >= 0 else "danger-inverse",
            ),
        ]

        for i, (label, value, style) in enumerate(cards):
            card = tb.Frame(card_row, bootstyle="secondary", padding=12)
            card.pack(side="left", expand=True, fill="x", padx=6)

            if i == 2:
                label_row = tb.Frame(card, bootstyle="secondary")
                label_row.pack(fill="x", anchor="w")
                
                tb.Label(label_row, text=label, font=("Helvetica", 9, "bold"),foreground= "#000000",
                         bootstyle="secondary-inverse").pack(side="left", anchor="s")
                
                tb.Label(label_row, text=f"({summary['status']})", font=("Helvetica", 9, "italic"), 
                         bootstyle=f"{style}-inverse").pack(side="left", anchor="s", padx=(10, 0))
                
                tb.Label(card, text=value, font=("Helvetica", 18, "bold"), foreground= "#000000", 
                         bootstyle=f"{style}-inverse").pack(anchor="w", pady=(2, 0))
            else: 
                tb.Label(card, text=label, font=("Helvetica", 9), foreground="#000000", bootstyle="secondary-inverse").pack(anchor="w")
                tb.Label(card, text=value, font=("Helvetica", 18, "bold"), foreground= "#000000", bootstyle=style).pack(anchor="w")


    def _build_action_bar(self):
        bar = tb.Frame(self)
        bar.pack(fill="x", padx=20, pady=(0, 4))

        tb.Label(bar, text="Double-click a cell to edit  •  Select a row and press Delete to remove",
             font=("Helvetica", 9, "bold"), bootstyle="secondary", foreground="#000000").pack(side="left")

        tb.Button(bar, text="🗑  Delete", bootstyle="outline-danger",
              command=self._on_delete).pack(side="right")
        tb.Button(bar, text="📤  Export", bootstyle="outline-success",
              command=self._on_export).pack(side="right", padx=(0, 6))


    def _build_table(self, transactions, summary):
        col_ids = [c[0] for c in self.COLUMNS]

        style_engine = self.winfo_toplevel().style
        style_engine.configure("secondary.Treeview", rowheight = 28)

        self._tree = tb.Treeview(
            self, columns=col_ids, show="headings",
            bootstyle="secondary", height=15, selectmode="browse",
        )

        widths = {"Date": 110, "Description": 200, "Category": 120,
                  "Type": 90, "Amount": 110}
        for col in col_ids:
            self._tree.heading(col, text=col)
            anchor = "e" if col == "Amount" else "w"
            self._tree.column(col, anchor=anchor, width=widths[col])

        self._tree.tag_configure("income",  foreground="#28a745")
        self._tree.tag_configure("expense", foreground="#dc3545")
        self._tree.tag_configure("net",     foreground="gray")

        if not transactions:
            self._tree.insert("", "end",
                              values=("—", "No transactions in this period", "", "", "—"),
                              tags=("net",))
        else:
            for row in transactions:
                tx_date, desc, cat, amount, tx_type= row
                signed = (f"+${amount:,.2f}" if tx_type == "Income"
                          else f"-${abs(amount):,.2f}")
                tag = "income" if tx_type == "Income" else "expense"
                self._tree.insert("", "end",
                                  values=(tx_date, desc, cat, tx_type, signed),
                                  tags=(tag,))

            net_str = (f"+${summary['net']:,.2f}" if summary["net"] >= 0
                       else f"-${abs(summary['net']):,.2f}")
            self._tree.insert("", "end",
                              values=(f"Net  ({summary['count']} transactions)",
                                      "", "", "", net_str),
                              tags=("net",))

        self._tree.bind("<Double-1>", self._on_double_click)
        self._tree.bind("<Delete>",   self._on_delete)

        scrollbar = tb.Scrollbar(self, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=scrollbar.set)

        self._tree.pack(side="left", fill=BOTH, expand=YES, padx=(20, 0), pady=10)
        scrollbar.pack(side="left", fill="y", pady=10, padx=(0, 20))


    def _on_double_click(self, event):
        item_id = self._tree.identify_row(event.y)
        col_id  = self._tree.identify_column(event.x)
        if not item_id or not col_id:
            return
        if "net" in self._tree.item(item_id, "tags"):
            return

        col_index   = int(col_id.replace("#", "")) - 1
        _, editable = self.COLUMNS[col_index]
        if not editable:
            return

        current_val = self._tree.item(item_id, "values")[col_index]
        if self.COLUMNS[col_index][0] == "Amount":
            current_val = current_val.replace("+", "").replace("-", "").replace("$", "").replace(",", "")

        self._editor = _CellEditor(
            tree=self._tree,
            item_id=item_id,
            col_index=col_index,
            current_value=current_val,
            on_commit=self._on_cell_commit,
        )

    def _on_cell_commit(self, item_id, col_index, new_value):
        values   = list(self._tree.item(item_id, "values"))
        col_name = self.COLUMNS[col_index][0]

        try:
            if col_name == "Date":
                date.fromisoformat(new_value)
            elif col_name == "Amount":
                amt = Decimal(new_value)
                if amt <= 0:
                    raise ValueError("Amount must be > 0")
            elif col_name == "Type":
                if new_value not in ("Income", "Expense"):
                    raise ValueError("Type must be Income or Expense")
        except (ValueError, InvalidOperation) as e:
            messagebox.showerror("Invalid Value", str(e))
            return

        if col_name == "Amount":
            tx_type = values[3]
            amt     = Decimal(new_value)
            values[col_index] = (f"+${amt:,.2f}" if tx_type == "Income"
                                 else f"-${amt:,.2f}")
        else:
            values[col_index] = new_value

        if col_name == "Type":
            tag = "income" if new_value == "Income" else "expense"
            self._tree.item(item_id, values=values, tags=(tag,))
        else:
            self._tree.item(item_id, values=values)

        self._save_row(item_id, values)

    def _save_row(self, item_id, values):
        all_items = self._tree.get_children()
        tx_items  = all_items[:-1]
        idx       = list(tx_items).index(item_id)
        raw       = self._raw_rows[idx]

        tx_id = _get_transaction_id(raw[0], raw[1], raw[3], raw[4])
        if tx_id is None:
            messagebox.showerror("Error", "Could not find transaction in database.")
            return

        try:
            new_date   = date.fromisoformat(str(values[0]))
            new_desc   = str(values[1]).strip()
            new_cat    = str(values[2]).strip()
            new_type   = str(values[3]).strip()
            new_amount = Decimal(
                str(values[4]).replace("+", "").replace("-", "")
                              .replace("$", "").replace(",", "")
            )
        except (ValueError, InvalidOperation) as e:
            messagebox.showerror("Parse Error", str(e))
            return

        success = _update_transaction(
            transaction_id=tx_id,
            tx_date=new_date,
            amount=new_amount,
            description=new_desc,
            tx_type=new_type,
            category_name=new_cat,
        )

        if success:
            self._raw_rows[idx] = (new_date, new_desc, new_cat,
                                   new_amount if new_type == "Income" else -new_amount,
                                   new_type)
        else:
            messagebox.showerror("Error", "Failed to save changes to database.")
            self._refresh()


    def _get_selected(self):
        selected = self._tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a row first.")
            return None, None

        item_id = selected[0]
        if "net" in self._tree.item(item_id, "tags"):
            messagebox.showwarning("Invalid Selection",
                                   "Please select a transaction row, not the net total.")
            return None, None

        all_items = self._tree.get_children()
        idx = list(all_items[:-1]).index(item_id)
        return item_id, self._raw_rows[idx]

    def _on_delete(self, _event=None):
        _, raw = self._get_selected()
        if raw is None:
            return

        confirmed = messagebox.askyesno(
            "Confirm Delete",
            f"Delete transaction permanently?\n\n"
            f"  Date: {raw[0]}\n"
            f"  Desc: {raw[1]}\n"
            f"  Amt:  ${abs(raw[3]):,.2f}\n\n"
            "This action cannot be undone.",
        )
        if not confirmed:
            return

        tx_id = _get_transaction_id(raw[0], raw[1], raw[3], raw[4])
        if tx_id is None:
            messagebox.showerror("Error", "Could not find transaction in database.")
            return

        if _delete_transaction(tx_id):
            self._refresh()
             
            # If the main app reference exists, call its load_graph method directly
            if self._app_ref and hasattr(self._app_ref, 'load_graph'):
                self._app_ref.load_graph()
        else:
            messagebox.showerror("Error", "Failed to delete transaction.")
    def _on_export(self):
        transactions, summary = get_transaction_ledger(
            preset=self._preset,
            cycle_start_day=self._cycle_start_day,
            custom_start=self._custom_start,
            custom_end=self._custom_end,
        )
        start, end = summary["start"], summary["end"]
        success = export_to_excel(transactions, summary, start, end)
        if success:
            messagebox.showinfo("Export Complete", "Transactions exported successfully.")        
